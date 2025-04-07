import os
import io
from time import sleep
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl import load_workbook


def create_excel_OLD(data_list, file_name, open_excel=False):
    """
    Create an xlsx file from a list of dictionaries.

    Args:
        data_list (list): List of dictionaries where each element represents a row and each dictionary key represents a column.
        file_name (str): Name of the xlsx file to create.
        open_excel (bool): If True, open the Excel file after creation.
    """
    # Create a new workbook.
    workbook = Workbook()

    # Select the active sheet (default is the first one).
    sheet = workbook.active

    # Set borders for all cells.
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border

    # Write headers (column names) in the first row and center-align the text.
    headers = list(data_list[0].keys())
    for idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=idx, value=header)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light gray background
        cell.font = Font(bold=True)  # Bold font
        cell.alignment = Alignment(horizontal='center')  # Center text
        cell.border = border  # Apply border to the cell

    # Write data in the remaining rows
    for row_idx, row_data in enumerate(data_list, start=2):
        for col_idx, value in enumerate(row_data.values(), start=1):

            # Try to convert to float if not already a float
            if not isinstance(value, float):
                try:
                    value = float(value)
                except:
                    pass

            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border  # Apply border to the cell

            # You can uncomment the line below to apply currency formatting (adjust as needed)
            # if isinstance(value, float):
            #     cell.number_format = '#,##0.00 [$€-C0A];[RED]-#,##0.00 [$€-C0A]'

    # Freeze the first row
    sheet.freeze_panes = "A2"

    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        for cell in column:
            max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # Save the workbook as an xlsx file
    workbook.save(file_name)

    # Automatically open the Excel file
    if open_excel:
        sleep(2)
        os.startfile(file_name)

def create_excel(data_list, file_name=None, open_excel=False, return_as_bytes=False, columns_order=None):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from time import sleep
    import os

    workbook = Workbook()
    sheet = workbook.active

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Determine column headers
    headers = columns_order or list(data_list[0].keys())

    # Write headers
    for idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=idx, value=header)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Write data
    for row_idx, row_data in enumerate(data_list, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = row_data.get(header)

            if isinstance(value, (list, dict)):
                value = str(value)
            elif not isinstance(value, (float, int, str, type(None))):
                value = str(value)
            else:
                if not isinstance(value, float):
                    try:
                        value = float(value)
                    except:
                        pass

            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

    sheet.freeze_panes = "A2"

    for column in sheet.columns:
        max_length = 0
        for cell in column:
            max_length = max(max_length, len(str(cell.value)) if cell.value else 0)
        adjusted_width = max_length + 2
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    if return_as_bytes:
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output
    else:
        if not file_name:
            raise ValueError("file_name must be provided unless return_as_bytes is True.")
        workbook.save(file_name)
        if open_excel:
            sleep(2)
            os.startfile(file_name)
        return None


def read_xlsx_to_dict(xlsx_file):
    """
    Read an xlsx file and return its content as a list of dictionaries.
    
    Args:
        xlsx_file (str): Path to the xlsx file.
    
    Returns:
        list: List of dictionaries representing the rows.
    """
    if not os.path.exists(xlsx_file):
        print(f"ERROR: The file {xlsx_file} does not exist!")
        input("pause ...")
        return []

    # Load the workbook
    wb = load_workbook(filename=xlsx_file)
    sheet = wb.active

    # Get the headers
    headers = [cell.value.lower() for cell in sheet[1]]

    # Get the data
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = {headers[i]: row[i] for i in range(len(headers))}
        data.append(row_dict)

    return data


def find_row_by_key(data, key, key_value):
    """
    Search for rows where the value in a specific key matches the given value.
    
    Args:
        data (list): List of dictionaries.
        key (str): Key to search by.
        key_value (Any): Value to match.
    
    Returns:
        list: Filtered list of dictionaries.
    """
    return [row for row in data if row[key] == key_value]


# Example usage:
# data = read_xlsx_to_dict('path/to/file.xlsx')
# print(data)
